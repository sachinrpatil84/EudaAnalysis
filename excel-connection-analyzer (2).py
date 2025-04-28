import os
import re
import pandas as pd
import xml.etree.ElementTree as ET
import zipfile
import openpyxl
from openpyxl.packaging.relationship import RelationshipList
import win32com.client
import csv
import logging
from pathlib import Path
import json
import sys
from typing import Dict, List, Tuple, Optional, Any, Union

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("ExcelConnectionAnalyzer")

class ConnectionAnalyzer:
    """
    A class to analyze data connections in Excel files.
    Supports XLSX, XLSM, and CSV files.
    """
    
    def __init__(self):
        self.connections = []
        self.complexity_thresholds = {
            "low": 3,     # Simple direct connections or single CSV imports
            "medium": 7,  # Multiple tables, simple transformations
            "high": 10    # Complex joins, custom SQL, multiple sources
        }
        
    def analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Main entry point to analyze an Excel or CSV file for connections
        
        Args:
            file_path: Path to the Excel or CSV file
            
        Returns:
            List of connections with their details
        """
        file_path = Path(file_path)
        logger.info(f"Analyzing file: {file_path}")
        
        if not file_path.exists():
            logger.error(f"File {file_path} does not exist")
            raise FileNotFoundError(f"File {file_path} does not exist")
        
        file_extension = file_path.suffix.lower()
        
        if file_extension in ['.xlsx', '.xlsm']:
            return self._analyze_excel_file(file_path)
        elif file_extension == '.csv':
            return [self._analyze_csv_file(file_path)]
        else:
            logger.error(f"Unsupported file format: {file_extension}")
            raise ValueError(f"Unsupported file format: {file_extension}")
    
    def _analyze_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Analyze an Excel file for various connection types
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of connections with their details
        """
        self.connections = []
        
        # Check if file can be opened with openpyxl
        try:
            # Extract and analyze internal connections using the ZIP structure
            self._extract_internal_connections(file_path)
            
            # Extract PowerQuery/Get & Transform connections
            self._extract_power_query_connections(file_path)
            
            # Use openpyxl for data connections and worksheets
            self._extract_openpyxl_connections(file_path)
            
            # If needed and on Windows, try to use COM objects for deeper analysis
            if os.name == 'nt' and not self.connections:
                try:
                    self._extract_com_connections(file_path)
                except Exception as e:
                    logger.warning(f"COM extraction failed: {e}")
        
        except Exception as e:
            logger.error(f"Error analyzing Excel file: {e}")
            
        # Remove duplicates based on connection names
        unique_connections = []
        connection_names = set()
        
        for conn in self.connections:
            if conn["Name"] not in connection_names:
                unique_connections.append(conn)
                connection_names.add(conn["Name"])
        
        return unique_connections
    
    def _extract_internal_connections(self, file_path: Path) -> None:
        """
        Extract connection information from the internal XML files in the Excel package
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                # Check for connections file
                if 'xl/connections.xml' in z.namelist():
                    with z.open('xl/connections.xml') as f:
                        root = ET.parse(f).getroot()
                        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        for connection in root.findall('.//x:connection', ns):
                            conn_name = connection.get('name', 'Unknown')
                            conn_type = connection.get('type', 'Unknown')
                            
                            # Get connection string if available
                            connection_string = None
                            dbPr = connection.find('.//x:dbPr', ns)
                            if dbPr is not None:
                                connection_string = dbPr.get('connection', '')
                            
                            # Get query text if available
                            query_text = None
                            command_text = connection.find('.//x:commandText', ns)
                            if command_text is not None:
                                query_text = command_text.text
                            
                            # Get target worksheets
                            worksheet_name = "Unknown"
                            tables = []
                            
                            # Process table data if present
                            table_refs = []
                            for table_ref in connection.findall('.//x:tableRef', ns):
                                table_refs.append(table_ref.get('id'))
                            
                            if table_refs:
                                # Look up table information to find worksheet
                                for table_id in table_refs:
                                    tables_path = f'xl/tables/table{table_id}.xml'
                                    if tables_path in z.namelist():
                                        with z.open(tables_path) as tf:
                                            table_root = ET.parse(tf).getroot()
                                            ref = table_root.get('ref', 'Unknown')
                                            if '!' in ref:
                                                worksheet_name = ref.split('!')[0].replace("'", "")
                                            tables.append(table_root.get('displayName', f'Table{table_id}'))
                            
                            # Create connection record
                            connection_info = {
                                "Name": conn_name,
                                "Type": self._determine_connection_type(conn_type, connection_string),
                                "Connection String": connection_string or "N/A",
                                "Target Datasources": self._extract_target_datasources(connection_string),
                                "Query Text": query_text or "N/A",
                                "Worksheet Name": worksheet_name,
                                "Tables": tables if tables else ["N/A"],
                                "Complexity Score": self._calculate_complexity_score(
                                    conn_type, connection_string, query_text
                                ),
                                "Purpose Description": self._infer_connection_purpose(
                                    conn_name, connection_string, query_text
                                )
                            }
                            
                            self.connections.append(connection_info)
                
                # Check for external links
                if 'xl/externalLinks/_rels' in z.namelist():
                    # Process external links
                    external_links = [f for f in z.namelist() if f.startswith('xl/externalLinks/') and f.endswith('.xml')]
                    
                    for link_file in external_links:
                        with z.open(link_file) as f:
                            link_root = ET.parse(f).getroot()
                            ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                            
                            # Get external book reference
                            ext_ref = link_root.find('.//x:externalBook', ns)
                            if ext_ref is not None:
                                # Get the relationship file to find the target
                                link_id = link_file.split('/')[-1].replace('.xml', '')
                                rel_file = f'xl/externalLinks/_rels/{link_id}.xml.rels'
                                
                                if rel_file in z.namelist():
                                    with z.open(rel_file) as rf:
                                        rel_root = ET.parse(rf).getroot()
                                        ns_rel = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                                        
                                        for rel in rel_root.findall('.//r:Relationship', ns_rel):
                                            target = rel.get('Target', '')
                                            
                                            connection_info = {
                                                "Name": f"External Link {link_id}",
                                                "Type": "External Workbook",
                                                "Connection String": target,
                                                "Target Datasources": [target],
                                                "Query Text": "N/A",
                                                "Worksheet Name": "Multiple",
                                                "Tables": ["N/A"],
                                                "Complexity Score": 4,  # Medium-low complexity
                                                "Purpose Description": f"External data reference to {target}"
                                            }
                                            
                                            self.connections.append(connection_info)
                
                # Look for data connections in content types
                if '[Content_Types].xml' in z.namelist():
                    with z.open('[Content_Types].xml') as f:
                        root = ET.parse(f).getroot()
                        ns = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
                        
                        # Look for query table definitions
                        for override in root.findall('.//ct:Override', ns):
                            if 'queryTable' in override.get('PartName', ''):
                                query_table_path = override.get('PartName').lstrip('/')
                                
                                if query_table_path in z.namelist():
                                    with z.open(query_table_path) as qtf:
                                        qt_root = ET.parse(qtf).getroot()
                                        ns_qt = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                        
                                        # Get connection ID
                                        connection_id = qt_root.find('.//x:queryTableRefresh', ns_qt)
                                        if connection_id is not None:
                                            conn_id = connection_id.get('connectionId')
                                            if conn_id:
                                                # Connection already processed through connections.xml
                                                continue
                
        except zipfile.BadZipFile:
            logger.error(f"File is not a valid Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Error extracting internal connections: {e}")
    
    def _extract_power_query_connections(self, file_path: Path) -> None:
        """
        Extract Power Query / Get & Transform connections from Excel file
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                # Check for Power Query connections
                if 'customXml/item1.xml' in z.namelist():
                    with z.open('customXml/item1.xml') as f:
                        try:
                            root = ET.parse(f).getroot()
                            
                            # Different namespaces depending on Excel version
                            namespaces = {
                                'pq': 'http://schemas.microsoft.com/DataMashup',
                                'dpm': 'http://schemas.microsoft.com/DataMashup',
                                'm': 'http://schemas.microsoft.com/DataMashup'
                            }
                            
                            # Try different namespaces
                            queries = None
                            for prefix, uri in namespaces.items():
                                queries = root.findall(f'.//{{{uri}}}Query')
                                if queries:
                                    break
                            
                            if not queries:
                                # Try without namespace
                                queries = root.findall('.//Query')
                            
                            if queries:
                                for idx, query in enumerate(queries):
                                    # Try to get name with different namespaces
                                    query_name = None
                                    for prefix, uri in namespaces.items():
                                        name_elem = query.find(f'.//{{{uri}}}Name')
                                        if name_elem is not None and name_elem.text:
                                            query_name = name_elem.text
                                            break
                                    
                                    if not query_name:
                                        # Try without namespace
                                        name_elem = query.find('.//Name')
                                        if name_elem is not None and name_elem.text:
                                            query_name = name_elem.text
                                    
                                    if not query_name:
                                        query_name = f"PowerQuery_{idx}"
                                    
                                    # Extract formula code if available
                                    query_formula = None
                                    for prefix, uri in namespaces.items():
                                        formula_elem = query.find(f'.//{{{uri}}}Formula')
                                        if formula_elem is not None and formula_elem.text:
                                            query_formula = formula_elem.text
                                            break
                                    
                                    if not query_formula:
                                        # Try without namespace
                                        formula_elem = query.find('.//Formula')
                                        if formula_elem is not None and formula_elem.text:
                                            query_formula = formula_elem.text
                                    
                                    # Source type detection
                                    source_type = "Power Query"
                                    target_datasources = ["Unknown"]
                                    
                                    if query_formula:
                                        # Try to extract source information from the formula
                                        if "Sql.Database" in query_formula:
                                            source_type = "SQL Database via Power Query"
                                            # Extract server and database if possible
                                            server_match = re.search(r'Sql\.Database\("([^"]+)"', query_formula)
                                            db_match = re.search(r'Sql\.Database\([^,]+,\s*"([^"]+)"', query_formula)
                                            
                                            if server_match:
                                                target_server = server_match.group(1)
                                                target_db = db_match.group(1) if db_match else "Unknown"
                                                target_datasources = [f"{target_server}/{target_db}"]
                                        
                                        elif "Excel.Workbook" in query_formula or "Excel.CurrentWorkbook" in query_formula:
                                            source_type = "Excel Workbook via Power Query"
                                            # Extract file path if possible
                                            file_match = re.search(r'Excel\.Workbook\(\s*File\.Contents\s*\(\s*"([^"]+)"\s*\)', query_formula)
                                            if file_match:
                                                target_datasources = [file_match.group(1)]
                                        
                                        elif "Csv.Document" in query_formula:
                                            source_type = "CSV via Power Query"
                                            # Extract file path if possible
                                            file_match = re.search(r'Csv\.Document\(\s*File\.Contents\s*\(\s*"([^"]+)"\s*\)', query_formula)
                                            if file_match:
                                                target_datasources = [file_match.group(1)]
                                        
                                        elif "Web.Contents" in query_formula:
                                            source_type = "Web API via Power Query"
                                            # Extract URL if possible
                                            url_match = re.search(r'Web\.Contents\s*\(\s*"([^"]+)"\s*\)', query_formula)
                                            if url_match:
                                                target_datasources = [url_match.group(1)]
                                    
                                    # Create connection record
                                    connection_info = {
                                        "Name": query_name,
                                        "Type": source_type,
                                        "Connection String": "N/A", 
                                        "Target Datasources": target_datasources,
                                        "Query Text": query_formula or "N/A",
                                        "Worksheet Name": self._find_query_target_worksheet(file_path, query_name),
                                        "Tables": ["N/A"],
                                        "Complexity Score": self._calculate_powerquery_complexity(query_formula or ""),
                                        "Purpose Description": self._infer_connection_purpose(
                                            query_name, "", query_formula
                                        )
                                    }
                                    
                                    self.connections.append(connection_info)
                        except ET.ParseError:
                            logger.warning("Power Query XML could not be parsed")
                
                # Check for DataModel connections (newer format)
                if 'customXml' in z.namelist():
                    datamodel_files = [f for f in z.namelist() if f.startswith('customXml/item') and f.endswith('.xml')]
                    
                    for dm_file in datamodel_files:
                        try:
                            with z.open(dm_file) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                
                                # Look for connection strings in the data model
                                conn_matches = re.findall(r'<[^>]*ConnectionString="([^"]+)"', content)
                                
                                for idx, conn_str in enumerate(conn_matches):
                                    connection_name = f"DataModel_{Path(dm_file).stem}_{idx}"
                                    
                                    # Determine connection type and target
                                    conn_type = "Data Model Connection"
                                    target = ["Unknown"]
                                    
                                    if "Provider=SQLOLEDB" in conn_str or "Provider=SQLNCLI" in conn_str:
                                        conn_type = "SQL Server via Data Model"
                                        
                                        # Extract server and database
                                        server_match = re.search(r'Data Source=([^;]+)', conn_str)
                                        db_match = re.search(r'Initial Catalog=([^;]+)', conn_str)
                                        
                                        if server_match:
                                            server = server_match.group(1)
                                            db = db_match.group(1) if db_match else "Unknown"
                                            target = [f"{server}/{db}"]
                                    
                                    elif "OLEDB;Provider=Microsoft.ACE.OLEDB" in conn_str:
                                        conn_type = "Access/Excel via Data Model"
                                        # Extract file path
                                        file_match = re.search(r'Data Source=([^;]+)', conn_str)
                                        if file_match:
                                            target = [file_match.group(1)]
                                    
                                    # Create connection record
                                    connection_info = {
                                        "Name": connection_name,
                                        "Type": conn_type,
                                        "Connection String": conn_str,
                                        "Target Datasources": target,
                                        "Query Text": "N/A",
                                        "Worksheet Name": "Data Model",
                                        "Tables": ["N/A"],
                                        "Complexity Score": self._calculate_complexity_score(
                                            conn_type, conn_str, ""
                                        ),
                                        "Purpose Description": self._infer_connection_purpose(
                                            connection_name, conn_str, ""
                                        )
                                    }
                                    
                                    self.connections.append(connection_info)
                        except Exception as e:
                            logger.warning(f"Could not process data model file {dm_file}: {e}")
        
        except Exception as e:
            logger.error(f"Error extracting Power Query connections: {e}")
    
    def _extract_openpyxl_connections(self, file_path: Path) -> None:
        """
        Use openpyxl to extract additional connection information
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_links=True)
            
            # Check for defined names that might indicate connections
            for name in workbook.defined_names.definedName:
                if '_xlnm._FilterDatabase' in name.name or 'xl_connect' in name.name.lower():
                    # This might be a connection or filter
                    connection_info = {
                        "Name": name.name,
                        "Type": "Named Range Connection",
                        "Connection String": str(name.value),
                        "Target Datasources": ["Internal"],
                        "Query Text": "N/A",
                        "Worksheet Name": name.attr_text if hasattr(name, 'attr_text') else "Unknown",
                        "Tables": ["N/A"],
                        "Complexity Score": 2,  # Low complexity
                        "Purpose Description": "Named range potentially used for data filtering or connection"
                    }
                    
                    self.connections.append(connection_info)
            
            # Look for worksheets that might contain external data ranges
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Look for tables in the worksheet
                tables = getattr(sheet, 'tables', {})
                for table_name, table in tables.items():
                    if hasattr(table, 'connectionId') and table.connectionId:
                        # This is a table connected to external data
                        connection_info = {
                            "Name": f"Table_{table_name}",
                            "Type": "Table Connection",
                            "Connection String": "N/A",
                            "Target Datasources": ["Unknown"],
                            "Query Text": "N/A",
                            "Worksheet Name": sheet_name,
                            "Tables": [table_name],
                            "Complexity Score": 3,  # Low-medium complexity
                            "Purpose Description": f"Data table in worksheet {sheet_name}"
                        }
                        
                        self.connections.append(connection_info)
            
            # Check for pivot tables
            for sheet_name in workbook.sheetnames:
                # Access pivot tables if the attribute exists
                sheet = workbook[sheet_name]
                pivot_tables = getattr(sheet, 'pivot_tables', [])
                
                for pt_idx, pivot_table in enumerate(pivot_tables):
                    cache_id = getattr(pivot_table, 'cache_id', None)
                    if cache_id:
                        connection_info = {
                            "Name": f"PivotTable_{sheet_name}_{pt_idx}",
                            "Type": "Pivot Table Connection",
                            "Connection String": f"Cache ID: {cache_id}",
                            "Target Datasources": ["Internal Data"],
                            "Query Text": "N/A",
                            "Worksheet Name": sheet_name,
                            "Tables": ["N/A"],
                            "Complexity Score": 5,  # Medium complexity
                            "Purpose Description": f"Pivot table in worksheet {sheet_name}"
                        }
                        
                        self.connections.append(connection_info)
            
        except Exception as e:
            logger.warning(f"Error in openpyxl analysis: {e}")
    
    def _extract_com_connections(self, file_path: Path) -> None:
        """
        Use COM objects (Windows only) to extract connections
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            try:
                workbook = excel.Workbooks.Open(str(file_path.absolute()))
                
                # Check for query tables
                for sheet in workbook.Sheets:
                    try:
                        # Check if the sheet has query tables
                        if hasattr(sheet, 'QueryTables'):
                            for qt in sheet.QueryTables:
                                connection_info = {
                                    "Name": qt.Name,
                                    "Type": "Query Table",
                                    "Connection String": qt.Connection,
                                    "Target Datasources": [self._extract_target_from_connection(qt.Connection)],
                                    "Query Text": qt.CommandText if hasattr(qt, 'CommandText') else "N/A",
                                    "Worksheet Name": sheet.Name,
                                    "Tables": ["N/A"],
                                    "Complexity Score": 5,  # Medium complexity
                                    "Purpose Description": self._infer_connection_purpose(
                                        qt.Name, qt.Connection, 
                                        qt.CommandText if hasattr(qt, 'CommandText') else ""
                                    )
                                }
                                
                                self.connections.append(connection_info)
                    except Exception as sheet_error:
                        logger.warning(f"Error processing sheet: {sheet_error}")
                
                # Check for connections in the workbook
                if hasattr(workbook, 'Connections'):
                    for conn in workbook.Connections:
                        try:
                            connection_info = {
                                "Name": conn.Name,
                                "Type": self._get_com_connection_type(conn),
                                "Connection String": conn.ODBCConnection.Connection if hasattr(conn, 'ODBCConnection') else "N/A",
                                "Target Datasources": [self._extract_target_from_connection(
                                    conn.ODBCConnection.Connection if hasattr(conn, 'ODBCConnection') else ""
                                )],
                                "Query Text": conn.ODBCConnection.CommandText if hasattr(conn, 'ODBCConnection') else "N/A",
                                "Worksheet Name": "Multiple",
                                "Tables": ["N/A"],
                                "Complexity Score": 6,  # Medium complexity
                                "Purpose Description": self._infer_connection_purpose(
                                    conn.Name,
                                    conn.ODBCConnection.Connection if hasattr(conn, 'ODBCConnection') else "",
                                    conn.ODBCConnection.CommandText if hasattr(conn, 'ODBCConnection') else ""
                                )
                            }
                            
                            self.connections.append(connection_info)
                        except Exception as conn_error:
                            logger.warning(f"Error processing connection: {conn_error}")
                
            finally:
                workbook.Close(False)
                excel.Quit()
        
        except Exception as e:
            logger.warning(f"COM extraction error: {e}")
    
    def _analyze_csv_file(self, file_path: Path) -> Dict[str, Any]:
        """
        Analyze a CSV file
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Dictionary with CSV connection information
        """
        try:
            # Read the first few lines to determine the structure
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                sample = ''.join([f.readline() for _ in range(5)])
            
            # Detect dialect
            dialect = csv.Sniffer().sniff(sample)
            delimiter = dialect.delimiter
            
            # Read sample data to infer purpose
            df = pd.read_csv(file_path, nrows=10)
            column_count = len(df.columns)
            
            # Determine complexity based on columns and structure
            complexity = 2  # Base score for CSV
            
            if column_count > 10:
                complexity += 1
            if column_count > 20:
                complexity += 1
            
            # Check for date columns which might indicate time series data
            date_columns = 0
            for col in df.columns:
                if 'date' in col.lower() or 'time' in col.lower():
                    date_columns += 1
                    
            if date_columns > 0:
                complexity += 1
            
            # Infer purpose based on column names
            purpose = self._infer_csv_purpose(df.columns)
            
            return {
                "Name": file_path.name,
                "Type": "CSV File",
                "Connection String": f"File={file_path}",
                "Target Datasources": [str(file_path)],
                "Query Text": "N/A",
                "Worksheet Name": "N/A",
                "Tables": ["N/A"],
                "Complexity Score": complexity,
                "Purpose Description": purpose
            }
            
        except Exception as e:
            logger.error(f"Error analyzing CSV file: {e}")
            return {
                "Name": file_path.name,
                "Type": "CSV File (Error)",
                "Connection String": f"File={file_path}",
                "Target Datasources": [str(file_path)],
                "Query Text": "N/A",
                "Worksheet Name": "N/A",
                "Tables": ["N/A"],
                "Complexity Score": 1,
                "Purpose Description": f"Error analyzing file: {str(e)}"
            }
    
    def _determine_connection_type(self, conn_type: str, conn_string: Optional[str]) -> str:
        """
        Determine the specific connection type based on connection string and type
        
        Args:
            conn_type: The connection type from Excel
            conn_string: The connection string
            
        Returns:
            A more specific connection type
        """
        if not conn_string:
            return conn_type
        
        conn_string = conn_string.lower()
        
        if "provider=sqloledb" in conn_string or "provider=sqlncli" in conn_string:
            return "SQL Server"
        elif "provider=microsoft.ace.oledb" in conn_string and ".csv" in conn_string:
            return "CSV via ACE OLEDB"
        elif "provider=microsoft.ace.oledb" in conn_string and ".xls" in conn_string:
            return "Excel via ACE OLEDB"
        elif "provider=microsoft.ace.oledb" in conn_string and ".accdb" in conn_string:
            return "Access Database"
        elif "dsn=" in conn_string:
            return "ODBC DSN"
        elif "oledb" in conn_string:
            return "OLEDB Generic"
        elif "web" in conn_type.lower() or "url" in conn_string:
            return "Web Query"
        else:
            return conn_type
    
    def _extract_target_datasources(self, conn_string: Optional[str]) -> List[str]:
        """
        Extract target data source names from connection string
        
        Args:
            conn_string: The connection string
            
        Returns:
            List of target data sources
        """
        if not conn_string:
            return ["Unknown"]
        
        # For SQL Server connections
        if "provider=sqloledb" in conn_string.lower() or "provider=sqlncli" in conn_string.lower():
            server = re.search(r'Data Source=([^;]+)', conn_string)
            db = re.search(r'Initial Catalog=([^;]+)', conn_string)
            
            if server:
                server_name = server.group(1)
                db_name = db.group(1) if db else "Unknown"
                return [f"{server_name}/{db_name}"]
        
        # For CSV or Excel files via ACE OLEDB
        if "provider=microsoft.ace.oledb" in conn_string.lower():
            file_path = re.search(r'Data Source=([^;]+)', conn_string)
            if file_path:
                return [file_path.group(1)]
        
        # For DSN connections
        if "dsn=" in conn_string.lower():
            dsn = re.search(r'DSN=([^;]+)', conn_string)
            if dsn:
                return [f"DSN:{dsn.group(1)}"]
        
        # For web queries
        if "url=" in conn_string.lower():
            url = re.search(r'URL=([^;]+)', conn_string)
            if url:
                return [url.group(1)]
        
        return ["Unknown"]
    
    def _extract_target_from_connection(self, conn_string: str) -> str:
        """
        Extract a single target from connection string
        
        Args:
            conn_string: The connection string
            
        Returns:
            Target data source name
        """
        targets = self._extract_target_datasources(conn_string)
        return targets[0] if targets else "Unknown"
    
    def _calculate_complexity_score(self, conn_type: str, conn_string: Optional[str], query_text: Optional[str]) -> int:
        """
        Calculate a complexity score based on connection details
        
        Args:
            conn_type: The connection type
            conn_string: The connection string
            query_text: The SQL or query text
            
        Returns:
            Complexity score (1-10)
        """
        score = 3  # Base score
        
        # Connection type complexity
        if "sql" in conn_type.lower():
            score += 2  # SQL connections are more complex
        elif "oledb" in conn_type.lower():
            score += 1
        elif "web" in conn_type.lower():
            score += 2  # Web queries can be complex
        elif "csv" in conn_type.lower():
            score += 0  # CSV is simple
        
        # Query complexity
        if query_text:
            # Count joins
            join_count = query_text.lower().count(" join ")
            score += min(join_count, 3)  # Max 3 points for joins
            
            # Check for subqueries or CTEs
            if "select" in query_text.lower() and "from" in query_text.lower():
                select_count = query_text.lower().count("select ")
                if select_count > 1:
                    score += 1  # Has subqueries
            
            if "with " in query_text.lower() and " as (" in query_text.lower():
                score += 1  # Has CTEs
            
            # Check for window functions
            if "over (" in query_text.lower():
                score += 1
        
        # Connection string complexity
        if conn_string:
            # Count parameters
            param_count = conn_string.count(";")
            score += min(param_count // 3, 2)  # Max 2 points for params
        
        # Cap the score at 10
        return min(score, 10)
    
    def _calculate_powerquery_complexity(self, formula: str) -> int:
        """
        Calculate a complexity score for Power Query formulas
        
        Args:
            formula: The Power Query M formula
            
        Returns:
            Complexity score (1-10)
        """
        if not formula:
            return 3  # Default medium-low
        
        score = 3  # Base score
        
        # Count steps (let statements)
        step_count = formula.count("    ")  # PowerQuery often indents steps
        score += min(step_count // 3, 3)  # Max 3 points for steps
        
        # Check for complex operations
        operations = [
            "Table.Join", "Table.NestedJoin", "Table.Combine", "Table.Pivot", 
            "Table.UnPivot", "Table.Group", "Table.TransformColumns", 
            "Table.FillDown", "Table.FillUp", "Table.Transpose"
        ]
        
        for op in operations:
            if op in formula:
                score += 1
                if score >= 10:
                    break
        
        # Check for custom functions
        if "let func =" in formula or "as function" in formula:
            score += 2
        
        # Cap the score at 10
        return min(score, 10)
    
    def _infer_connection_purpose(self, conn_name: str, conn_string: Optional[str], query_text: Optional[str]) -> str:
        """
        Infer the purpose of a connection based on its properties
        
        Args:
            conn_name: The connection name
            conn_string: The connection string
            query_text: The SQL or query text
            
        Returns:
            Inferred purpose description
        """
        purpose = "Data connection"
        
        # Check connection name for clues
        conn_name_lower = conn_name.lower()
        
        if any(term in conn_name_lower for term in ["sales", "revenue", "income"]):
            purpose = "Sales or revenue data connection"
        elif any(term in conn_name_lower for term in ["customer", "client", "account"]):
            purpose = "Customer data connection"
        elif any(term in conn_name_lower for term in ["product", "inventory", "stock"]):
            purpose = "Product or inventory data connection"
        elif any(term in conn_name_lower for term in ["employee", "staff", "hr", "personnel"]):
            purpose = "Employee or HR data connection"
        elif any(term in conn_name_lower for term in ["finance", "accounting", "budget"]):
            purpose = "Financial data connection"
        elif any(term in conn_name_lower for term in ["report", "dashboard"]):
            purpose = "Reporting data connection"
        
        # Check query text for more clues
        if query_text:
            query_lower = query_text.lower()
            
            # Check for aggregations
            if any(agg in query_lower for agg in ["sum(", "count(", "avg(", "group by"]):
                purpose += " for aggregated data analysis"
            
            # Check for specific tables in FROM clauses
            from_tables = []
            from_matches = re.finditer(r'from\s+([a-zA-Z0-9_\.]+)', query_lower)
            for match in from_matches:
                table_name = match.group(1).split('.')[-1]
                from_tables.append(table_name)
            
            if from_tables:
                table_purpose = self._infer_purpose_from_table_names(from_tables)
                if table_purpose:
                    purpose = table_purpose
        
        return purpose
    
    def _infer_purpose_from_table_names(self, table_names: List[str]) -> str:
        """
        Infer purpose from table names
        
        Args:
            table_names: List of table names
            
        Returns:
            Inferred purpose description
        """
        table_purposes = {}
        
        for table in table_names:
            table_lower = table.lower()
            
            if any(term in table_lower for term in ["sale", "order", "invoice", "revenue"]):
                table_purposes["sales"] = table_purposes.get("sales", 0) + 1
            elif any(term in table_lower for term in ["customer", "client", "account"]):
                table_purposes["customer"] = table_purposes.get("customer", 0) + 1
            elif any(term in table_lower for term in ["product", "item", "sku", "inventory"]):
                table_purposes["product"] = table_purposes.get("product", 0) + 1
            elif any(term in table_lower for term in ["employee", "staff", "personnel"]):
                table_purposes["employee"] = table_purposes.get("employee", 0) + 1
            elif any(term in table_lower for term in ["finance", "gl", "ledger", "accounting"]):
                table_purposes["finance"] = table_purposes.get("finance", 0) + 1
        
        if not table_purposes:
            return ""
        
        # Find the most common purpose
        max_purpose = max(table_purposes.items(), key=lambda x: x[1])
        
        purpose_map = {
            "sales": "Sales or order data connection",
            "customer": "Customer data connection",
            "product": "Product or inventory data connection",
            "employee": "Employee or HR data connection",
            "finance": "Financial data connection"
        }
        
        return purpose_map.get(max_purpose[0], "")
    
    def _infer_csv_purpose(self, column_names: List[str]) -> str:
        """
        Infer the purpose of a CSV file based on column names
        
        Args:
            column_names: List of column names
            
        Returns:
            Inferred purpose description
        """
        purpose_indicators = {
            "sales": ["sale", "order", "invoice", "revenue", "transaction"],
            "customer": ["customer", "client", "account", "contact", "lead"],
            "product": ["product", "item", "sku", "inventory", "stock"],
            "employee": ["employee", "staff", "personnel", "hr", "payroll"],
            "finance": ["finance", "account", "budget", "cost", "expense", "profit"],
            "time_series": ["date", "time", "period", "month", "year", "quarter"],
            "geographic": ["country", "state", "city", "zip", "postal", "region"]
        }
        
        # Convert column names to lowercase for matching
        columns_lower = [str(col).lower() for col in column_names]
        
        # Count matches for each purpose
        purpose_counts = {purpose: 0 for purpose in purpose_indicators}
        
        for col in columns_lower:
            for purpose, indicators in purpose_indicators.items():
                if any(indicator in col for indicator in indicators):
                    purpose_counts[purpose] += 1
        
        # Find the most likely purpose
        max_count = max(purpose_counts.values())
        if max_count == 0:
            return "General data file"
        
        top_purposes = [p for p, count in purpose_counts.items() if count == max_count]
        
        purpose_descriptions = {
            "sales": "Sales or order data",
            "customer": "Customer or client data",
            "product": "Product or inventory data",
            "employee": "Employee or HR data",
            "finance": "Financial data",
            "time_series": "Time series data",
            "geographic": "Geographic or regional data"
        }
        
        if len(top_purposes) == 1:
            return purpose_descriptions[top_purposes[0]]
        elif "time_series" in top_purposes:
            # If time series is one of the top, combine it with the other purpose
            other_purposes = [p for p in top_purposes if p != "time_series"]
            if other_purposes:
                return f"Time series {purpose_descriptions[other_purposes[0]].lower()}"
            else:
                return purpose_descriptions["time_series"]
        else:
            # Combine the top two purposes
            if len(top_purposes) >= 2:
                return f"{purpose_descriptions[top_purposes[0]]} with {purpose_descriptions[top_purposes[1]].lower()}"
            else:
                return purpose_descriptions[top_purposes[0]]
    
    def _find_query_target_worksheet(self, file_path: Path, query_name: str) -> str:
        """
        Try to find which worksheet a Power Query feeds into
        
        Args:
            file_path: Path to the Excel file
            query_name: Name of the Power Query
            
        Returns:
            Worksheet name or "Unknown"
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Check if the query name matches a table name
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Check for tables with matching name
                tables = getattr(sheet, 'tables', {})
                for table_name in tables:
                    if query_name.lower() == table_name.lower():
                        return sheet_name
            
            return "Unknown"
        except Exception:
            return "Unknown"
    
    def _get_com_connection_type(self, connection) -> str:
        """
        Get connection type from COM object
        
        Args:
            connection: COM connection object
            
        Returns:
            Connection type string
        """
        try:
            if hasattr(connection, 'Type'):
                type_map = {
                    1: "ODBC Connection",
                    2: "DAO Connection",
                    3: "Web Query",
                    4: "OLE DB Connection",
                    5: "Text Connection",
                    6: "ADO Connection",
                    7: "DSN Connection",
                    8: "Office Data Connection"
                }
                return type_map.get(connection.Type, f"Connection Type {connection.Type}")
            
            if hasattr(connection, 'ODBCConnection'):
                return "ODBC Connection"
            elif hasattr(connection, 'OLEDBConnection'):
                return "OLE DB Connection"
            else:
                return "Unknown Connection"
        except Exception:
            return "Connection"

def analyze_excel_connections(file_path: str) -> List[Dict[str, Any]]:
    """
    Analyze an Excel file for data connections
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of dictionaries with connection information
    """
    analyzer = ConnectionAnalyzer()
    return analyzer.analyze_file(file_path)

def display_results(connections: List[Dict[str, Any]]) -> None:
    """
    Display connection analysis results in a formatted table
    
    Args:
        connections: List of connections with their details
    """
    if not connections:
        print("No connections found in the file.")
        return
    
    print(f"\n{'='*80}")
    print(f"{'CONNECTION ANALYSIS RESULTS':^80}")
    print(f"{'='*80}")
    
    for idx, conn in enumerate(connections, 1):
        print(f"\nConnection #{idx}: {conn['Name']}")
        print(f"{'-'*80}")
        print(f"Type: {conn['Type']}")
        print(f"Connection String: {conn['Connection String']}")
        print(f"Target Datasources: {', '.join(conn['Target Datasources'])}")
        print(f"Worksheet Name: {conn['Worksheet Name']}")
        
        if conn['Query Text'] != "N/A" and len(conn['Query Text']) > 100:
            print(f"Query Text: {conn['Query Text'][:100]}...")
        else:
            print(f"Query Text: {conn['Query Text']}")
            
        print(f"Complexity Score: {conn['Complexity Score']}/10")
        print(f"Purpose: {conn['Purpose Description']}")

def export_to_csv(connections: List[Dict[str, Any]], output_file: str) -> None:
    """
    Export connection analysis results to a CSV file
    
    Args:
        connections: List of connections with their details
        output_file: Path to the output CSV file
    """
    if not connections:
        print("No connections found to export.")
        return
    
    # Flatten target datasources and tables for CSV export
    for conn in connections:
        conn['Target Datasources'] = ', '.join(conn['Target Datasources'])
        if 'Tables' in conn:
            conn['Tables'] = ', '.join(conn['Tables'])
    
    # Convert to DataFrame for easy CSV export
    df = pd.DataFrame(connections)
    
    # Reorder columns for better readability
    columns = [
        'Name', 'Type', 'Connection String', 'Target Datasources', 
        'Query Text', 'Worksheet Name', 'Tables', 'Complexity Score', 
        'Purpose Description'
    ]
    
    # Only include columns that exist
    columns = [col for col in columns if col in df.columns]
    
    df = df[columns]
    df.to_csv(output_file, index=False)
    print(f"\nResults exported to {output_file}")

def export_to_json(connections: List[Dict[str, Any]], output_file: str) -> None:
    """
    Export connection analysis results to a JSON file
    
    Args:
        connections: List of connections with their details
        output_file: Path to the output JSON file
    """
    if not connections:
        print("No connections found to export.")
        return
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(connections, f, indent=2)
    
    print(f"\nResults exported to {output_file}")

def main():
    """
    Main function to run the Excel Connection Analyzer
    """
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel Connection Analyzer')
    parser.add_argument('file', help='Path to the Excel or CSV file')
    parser.add_argument('--output', '-o', help='Output file path for CSV or JSON export')
    parser.add_argument('--format', '-f', choices=['csv', 'json'], default='csv', 
                        help='Output format (csv or json)')
    
    args = parser.parse_args()
    
    try:
        # Analyze the file
        connections = analyze_excel_connections(args.file)
        
        # Display results
        display_results(connections)
        
        # Export results if output file specified
        if args.output:
            if args.format == 'csv':
                export_to_csv(connections, args.output)
            else:
                export_to_json(connections, args.output)
    
    except Exception as e:
        logger.error(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
