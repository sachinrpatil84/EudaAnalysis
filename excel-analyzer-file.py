import os
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from io import BytesIO

class ExcelAnalyzer:
    def __init__(self):
        """Initialize Excel analyzer"""
        self.file_path = None
        self.workbook = None
        self.metadata = {}
    
    def load_file(self, file_path):
        """Load Excel file for analysis"""
        try:
            self.file_path = file_path
            self.workbook = load_workbook(file_path, data_only=False)
            print(f"Excel file loaded: {file_path}")
            return True
        except Exception as e:
            print(f"Error loading Excel file: {str(e)}")
            return False
    
    def analyze_file(self):
        """Analyze Excel file and extract metadata"""
        if not self.workbook:
            print("No Excel file loaded.")
            return None
        
        try:
            # Basic file info
            file_name = os.path.basename(self.file_path)
            file_size = os.path.getsize(self.file_path) / 1024  # Size in KB
            sheet_count = len(self.workbook.sheetnames)
            
            # Initialize metadata
            self.metadata = {
                "filename": file_name,
                "file_size_kb": round(file_size, 2),
                "sheet_count": sheet_count,
                "sheets": self.workbook.sheetnames,
                "has_macros": self.check_for_macros(),
                "formulas": self.extract_formulas(),
                "external_connections": self.check_external_connections(),
                "vlookups_count": self.count_vlookups(),
                "pivot_tables": self.check_pivot_tables(),
                "charts": self.check_charts(),
                "cell_count": self.count_cells(),
                "data_validation": self.check_data_validation(),
                "conditional_formatting": self.check_conditional_formatting()
            }
            
            return self.metadata
        except Exception as e:
            print(f"Error analyzing Excel file: {str(e)}")
            return None

    def check_for_macros(self):
        """Check if the Excel file contains macros/VBA"""
        try:
            # Check if the workbook has VBA modules
            return self.workbook.vba_archive is not None
        except:
            # If vba_archive attribute is not available, check file extension
            return self.file_path.endswith('.xlsm') or self.file_path.endswith('.xls')
    
    def extract_formulas(self):
        """Extract formulas used in the workbook"""
        formula_info = {
            "total_count": 0,
            "by_type": {},
            "complex_formulas": [],
            "samples": []
        }
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.formula:
                            formula_info["total_count"] += 1
                            
                            # Categorize by formula type
                            formula_type = self._get_formula_type(cell.formula)
                            if formula_type in formula_info["by_type"]:
                                formula_info["by_type"][formula_type] += 1
                            else:
                                formula_info["by_type"][formula_type] = 1
                            
                            # Collect complex formulas (nested or long)
                            if len(cell.formula) > 50 or cell.formula.count('(') > 2:
                                if len(formula_info["complex_formulas"]) < 10:  # Limit to 10 examples
                                    formula_info["complex_formulas"].append({
                                        "sheet": sheet_name,
                                        "cell": cell.coordinate,
                                        "formula": cell.formula
                                    })
                            
                            # Collect formula samples
                            if len(formula_info["samples"]) < 20:  # Limit to 20 samples
                                formula_info["samples"].append({
                                    "sheet": sheet_name,
                                    "cell": cell.coordinate,
                                    "formula": cell.formula
                                })
                                
            return formula_info
        except Exception as e:
            print(f"Error extracting formulas: {str(e)}")
            return formula_info
    
    def _get_formula_type(self, formula):
        """Determine the formula type based on the function name"""
        common_functions = [
            "SUM", "AVERAGE", "COUNT", "MAX", "MIN", "IF", "VLOOKUP", "HLOOKUP", 
            "INDEX", "MATCH", "SUMIF", "COUNTIF", "SUMIFS", "COUNTIFS", "IFERROR",
            "CONCATENATE", "CONCAT", "OFFSET", "INDIRECT", "LOOKUP", "CHOOSE"
        ]
        
        for func in common_functions:
            if func in formula.upper():
                return func
        
        return "OTHER"
    
    def check_external_connections(self):
        """Check for external data connections"""
        connections = {
            "has_connections": False,
            "connection_types": []
        }
        
        # Connection indicators in formulas
        connection_indicators = [
            "ODBC", "JDBC", "SQL", "CONNECTION", "CUBEMEMBER", "CUBEVALUE", 
            "CUBESET", "OLAP", "WEBSERVICE", "FILTERXML", "WEBQUERY"
        ]
        
        try:
            # Check for connection indicators in formulas
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.formula:
                            for indicator in connection_indicators:
                                if indicator in cell.formula.upper():
                                    connections["has_connections"] = True
                                    if indicator not in connections["connection_types"]:
                                        connections["connection_types"].append(indicator)
            
            # Check for connection types by file extension
            if self.file_path.endswith('.xlsb'):
                connections["has_connections"] = True
                if "BINARY" not in connections["connection_types"]:
                    connections["connection_types"].append("BINARY")
                    
            return connections
        except Exception as e:
            print(f"Error checking external connections: {str(e)}")
            return connections
    
    def count_vlookups(self):
        """Count VLOOKUP formulas in the workbook"""
        vlookup_count = 0
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.formula and "VLOOKUP" in cell.formula.upper():
                            vlookup_count += 1
                            
            return vlookup_count
        except Exception as e:
            print(f"Error counting VLOOKUPs: {str(e)}")
            return vlookup_count
    
    def check_pivot_tables(self):
        """Check for pivot tables in the workbook"""
        pivot_count = 0
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                # Look for pivot cache in the sheet properties
                if hasattr(sheet, '_pivots') and sheet._pivots:
                    pivot_count += len(sheet._pivots)
                
                # Alternative detection method - checking for typical pivot table formulas
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.formula and "GETPIVOTDATA" in cell.formula.upper():
                            pivot_count += 1
                            break
                            
            return pivot_count
        except Exception as e:
            print(f"Error checking pivot tables: {str(e)}")
            return pivot_count
    
    def check_charts(self):
        """Check for charts in the workbook"""
        chart_count = 0
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                if hasattr(sheet, '_charts') and sheet._charts:
                    chart_count += len(sheet._charts)
                    
            return chart_count
        except Exception as e:
            print(f"Error checking charts: {str(e)}")
            return chart_count
    
    def count_cells(self):
        """Count cells with content in the workbook"""
        cell_counts = {
            "total": 0,
            "by_sheet": {}
        }
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet_cell_count = 0
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            sheet_cell_count += 1
                
                cell_counts["total"] += sheet_cell_count
                cell_counts["by_sheet"][sheet_name] = sheet_cell_count
                
            return cell_counts
        except Exception as e:
            print(f"Error counting cells: {str(e)}")
            return cell_counts
    
    def check_data_validation(self):
        """Check for data validation rules in the workbook"""
        validation_count = 0
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                if hasattr(sheet, 'data_validations') and sheet.data_validations:
                    validation_count += len(sheet.data_validations.dataValidation)
                    
            return validation_count
        except Exception as e:
            print(f"Error checking data validation: {str(e)}")
            return validation_count
    
    def check_conditional_formatting(self):
        """Check for conditional formatting in the workbook"""
        formatting_count = 0
        
        try:
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
                    formatting_count += len(sheet.conditional_formatting.cf_rules)
                    
            return formatting_count
        except Exception as e:
            print(f"Error checking conditional formatting: {str(e)}")
            return formatting_count

    def extract_text_for_embedding(self):
        """Extract textual content from the Excel file for embedding"""
        if not self.workbook:
            print("No Excel file loaded.")
            return []
        
        text_chunks = []
        
        try:
            # Process each sheet
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                # Sheet content as text
                sheet_content = f"Sheet: {sheet_name}\n"
                
                # Extract formulas
                formulas_text = "Formulas:\n"
                has_formulas = False
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.formula:
                            has_formulas = True
                            formulas_text += f"Cell {cell.coordinate}: {cell.formula}\n"
                
                if has_formulas:
                    sheet_content += formulas_text
                
                # Extract cell values (limit to first 1000 non-empty cells)
                cell_count = 0
                values_text = "Values:\n"
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and cell_count < 1000:
                            cell_count += 1
                            values_text += f"Cell {cell.coordinate}: {str(cell.value)}\n"
                
                sheet_content += values_text
                
                # Add sheet content as a chunk
                text_chunks.append({
                    "type": "sheet_content",
                    "sheet": sheet_name,
                    "content": sheet_content
                })
                
            # General workbook properties
            properties_text = f"""
            File name: {os.path.basename(self.file_path)}
            Number of sheets: {len(self.workbook.sheetnames)}
            Sheets: {', '.join(self.workbook.sheetnames)}
            Has macros: {self.check_for_macros()}
            VLOOKUPs: {self.count_vlookups()}
            External connections: {self.check_external_connections()['has_connections']}
            """
            
            text_chunks.append({
                "type": "workbook_properties",
                "content": properties_text
            })
                
            return text_chunks
        except Exception as e:
            print(f"Error extracting text for embedding: {str(e)}")
            return []
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            print("Workbook closed.")
